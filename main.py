from aiohttp import ClientConnectorError
from openpyxl import load_workbook
from PIL import Image
from PIL import UnidentifiedImageError
import os
import asyncio
import aiohttp
import aiofiles
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive

gauth = GoogleAuth()
gauth.LocalWebserverAuth()


prefix = input("Введите префикс: ")
file_name = input("Введите название файла: ")
sheet_name = input("Введите название таблицы: ")
name_index = input("Введите индекс наименования товара: ")
hatch_index = input("Введите индекс штрих кода товара: ")
url_index = input("Введите индекс url товара: ")
IMAGES_FOLDER = f'{file_name} folder'
os.mkdir(IMAGES_FOLDER)
FILES_PATH = os.path.join(IMAGES_FOLDER, '')


def parse_excel(filename, sheet_name, name_index, hatch_index, url_index):
    wb = load_workbook(filename)
    sheet = wb[sheet_name]
    index = 2
    mylist = []
    while True:
        data = []
        res_name = sheet[f'{name_index}{index}'].value
        try:
            name = res_name.replace('"', '').replace('/', ' ').replace(':', '').replace('«', '')\
                .replace('»', '').replace('?','').replace('|', '').replace('<', '').replace('>', '')
        except Exception:
            name = res_name
        try:
            hatch = int(sheet[f'{hatch_index}{index}'].value)
        except Exception:
            hatch = sheet[f'{hatch_index}{index}'].value
        url = sheet[f'{url_index}{index}'].value
        url_for = url
        print(url_for)
        url_column_number = str(url_index) + str(index)
        if hatch is None and name is None and url is None:
            break
        if hatch is not None and not isinstance(hatch, str):
            data.append(int(hatch))
        else:
            data.append(hatch)
        data.append(url)
        data.append(name)
        data.append(url_column_number)
        mylist.append(data)
        index += 1
    return mylist


def transformation_image(file_path):
    try:
        old_img = Image.open(file_path)
        height = old_img.size[0]
        width = old_img.size[1]

        if height != width and width != height:
            if height < 1000 and width < 1000:
                fixed_width = 1000
                size = 500
            elif height < 1500 and width < 1500:
                fixed_width = 1500
                size = 750
            elif height < 2000 and width < 2000:
                fixed_width = 2000
                size = 1000
            elif height < 2500 and width < 2500:
                fixed_width = 2500
                size = 1250
            elif height < 3000 and width < 3000:
                fixed_width = 3000
                size = 1500
            elif height < 4000 and width < 4000:
                fixed_width = 4000
                size = 2000
            else:
                fixed_width = 5000
                size = 2500
            height2 = height // 2
            width2 = width // 2
            new_image = Image.new(old_img.mode,
                                  (fixed_width, fixed_width), 'white')
            new_image.paste(old_img, (size - height2, size - width2))
            new_image.save(file_path)
        else:
            return "Картинка уже квадратная"
    except UnidentifiedImageError:
        pass


async def download_image(data):
    a = 0
    sema = asyncio.BoundedSemaphore(5)
    while True:
        hatch, url, name, url_column_number = data[a][0], data[a][1], data[a][2], data[a][3]
        if url is None:
            a += 1
            continue
        if hatch is None or isinstance(hatch, str):
            hatch = name
        print(hatch)
        a += 1
        try:
            timeout = aiohttp.ClientTimeout(total=5)
            async with sema, aiohttp.ClientSession() as session:
                async with session.get(url, timeout=timeout, ssl=False) as resp:
                    if resp.status == 200:
                        r = await resp.read()
                        async with aiofiles.open(
                                f'{FILES_PATH}{str(hatch)}{prefix}.png', "wb"
                        ) as outfile:
                            await outfile.write(r)
                            transformation_image(f'{FILES_PATH}{str(hatch)}{prefix}.png')
                            image_validator(url_column_number, file_name, sheet_name, hatch, resp.status)
                    else:
                        image_validator(url_column_number, file_name, sheet_name, hatch, resp.status)
                        continue
        except (asyncio.TimeoutError, ClientConnectorError):
            image_validator(url_column_number, file_name, sheet_name, hatch, resp.status)
            continue


def image_validator(url_column_number, filename, sheet_name, hatch, status):
    wb = load_workbook(filename)
    sheet = wb[sheet_name]
    try:
        v_image = Image.open((f'{FILES_PATH}{str(hatch)}{prefix}.png'))
        v_image.verify()
        v_image.close()
    except Exception:
        url = sheet[url_column_number].value
        if status == 444:
            sheet[url_column_number] = url + "(Сервер блокирует скачивание)"
        else:
            sheet[url_column_number] = url + "(Не валидная ccылка)"
        wb.save(filename)
        wb.close()
        try:
            os.remove(f'{FILES_PATH}{str(hatch)}{prefix}.png')
        except Exception:
            pass


# def upload_dir(dir_path='', real_folder_id=''):
#     try:
#         drive = GoogleDrive(gauth)
#
#         for file_name in os.listdir(dir_path):
#             fileID = real_folder_id
#             my_file = drive.CreateFile({"title": f'{file_name}', "parents": [{"kind": "drive#fileLink", "id": fileID}]})
#             my_file.SetContentFile(os.path.join(dir_path, file_name))
#             my_file.Upload()
#             print(f'File {file_name} was uploaded!')
#         return 'Success!Have a good day!'
#     except Exception as _ex:
#         return 'Got some trouble, check your code please!'
#
#
# def main():
#     print(upload_dir(dir_path=IMAGES_FOLDER, real_folder_id='15dMpzAou2teJ2qmq-76ZHr4t9ujNIa9_'))



if __name__ == '__main__':
    p = parse_excel(file_name, sheet_name, name_index, hatch_index, url_index)
    print(p)
    loop = asyncio.get_event_loop()
    tasks = [loop.create_task(download_image(p))]
    loop.run_until_complete(asyncio.wait(tasks))
    loop.close()

